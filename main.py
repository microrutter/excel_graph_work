import sys
import openpyxl
from openpyxl.comments import Comment
from openpyxl.chart import BarChart, Reference


def get_workbook():
    workbook = None
    if len(sys.argv) > 1:
        workbook = openpyxl.load_workbook(sys.argv[1])
        return workbook
    print("Not enough arguments please  give path to excel sheet")
    return workbook


def sort_student_scores(max_column: int, max_row: int, sheet):
    sheet.cell(column=max_column + 1, row=1).value = "Student Total"
    for i in range(2, max_row + 1):
        count: int = 0
        for k in range(2, max_column + 1):
            if sheet.cell(column=k, row=i).value:
                count = count + sheet.cell(column=k, row=i).value
        sheet.cell(column=max_column + 1, row=i).value = count


def sort_question_scores(max_column: int, max_row: int, sheet):
    sheet.cell(column=1, row=max_row + 1).value = "Question Total"
    for i in range(2, max_column + 1):
        count: int = 0
        for k in range(2, max_row + 1):
            if sheet.cell(column=i, row=k).value:
                count = count + sheet.cell(column=i, row=k).value
        sheet.cell(column=i, row=max_row + 1).value = count


def sort_mean_top_bottom(sheet) -> dict:
    max_row = sheet.max_row
    max_column = sheet.max_column
    count: int = 0
    for row in range(2, max_row + 1):
        if sheet.cell(column=2, row=row).value:
            count = count + sheet.cell(column=2, row=row).value
    mean: float = count / (max_row - 1)
    top_count: int = 0
    top_mean: float = 0
    bottom_count: int = 0
    bottom_mean: float = 0
    for row in range(2, max_row + 1):
        value = sheet.cell(column=2, row=row).value
        if value:
            if value > mean:
                top_mean = top_mean + value
                top_count = top_count + 1
            else:
                bottom_mean = bottom_mean + value
                bottom_count = bottom_count + 1
    top_mean = top_mean / top_count
    bottom_mean = bottom_mean / bottom_count
    return {"mean": mean, "top_mean": top_mean, "bottom_mean": bottom_mean}


def sort_mean_question(mean: float, top_mean: float, bottom_mean: float, sheet) -> int:
    max_row = sheet.max_row
    max_column = sheet.max_column
    sheet.cell(column=max_column + 2, row=1).value = "High Quarter Question"
    sheet.cell(column=max_column + 3, row=1).value = "Score"
    sheet.cell(column=max_column + 6, row=1).value = "Mid High Quarter Question"
    sheet.cell(column=max_column + 7, row=1).value = "Score"
    sheet.cell(column=max_column + 10, row=1).value = "Mid Low Quarter Question"
    sheet.cell(column=max_column + 11, row=1).value = "Score"
    sheet.cell(column=max_column + 14, row=1).value = "Low Quarter Question"
    sheet.cell(column=max_column + 15, row=1).value = "Score"
    high: int = 2
    midhigh: int = 2
    midlow: int = 2
    low: int = 2
    for row in range(2, max_row + 1):
        value = sheet.cell(column=2, row=row).value
        if value >= top_mean:
            sheet.cell(column=max_column + 2, row=high).value = sheet.cell(
                column=1, row=row
            ).value
            sheet.cell(column=max_column + 3, row=high).value = value
            high = high + 1
        elif top_mean > value >= mean:
            sheet.cell(column=max_column + 6, row=midhigh).value = sheet.cell(
                column=1, row=row
            ).value
            sheet.cell(column=max_column + 7, row=midhigh).value = value
            midhigh = midhigh + 1
        elif mean > value >= bottom_mean:
            sheet.cell(column=max_column + 10, row=midlow).value = sheet.cell(
                column=1, row=row
            ).value
            sheet.cell(column=max_column + 11, row=midlow).value = value
            midlow = midlow + 1
        else:
            sheet.cell(column=max_column + 14, row=low).value = sheet.cell(
                column=1, row=row
            ).value
            sheet.cell(column=max_column + 15, row=low).value = value
            low = low + 1
    return max([high, midhigh, midlow, low])


def sort_mean_student(
    mean: float, top_mean: float, bottom_mean: float, sheet, max_row, max_row_results
):
    print(top_mean)
    print(mean)
    print(bottom_mean)
    sheet.cell(column=4, row=max_row + 2).value = "High Quarter Students"
    sheet.cell(column=5, row=max_row + 2).value = "Score"
    sheet.cell(column=8, row=max_row + 2).value = "Mid High Quarter Students"
    sheet.cell(column=9, row=max_row + 2).value = "Score"
    sheet.cell(column=12, row=max_row + 2).value = "Mid Low Quarter Students"
    sheet.cell(column=13, row=max_row + 2).value = "Score"
    sheet.cell(column=16, row=max_row + 2).value = "Low Quarter Students"
    sheet.cell(column=17, row=max_row + 2).value = "Score"
    high: int = max_row + 3
    midhigh: int = max_row + 3
    midlow: int = max_row + 3
    low: int = max_row + 3
    for row in range(2, max_row_results + 1):
        value = sheet.cell(column=2, row=row).value
        if value >= top_mean:
            sheet.cell(column=4, row=high).value = sheet.cell(column=1, row=row).value
            sheet.cell(column=5, row=high).value = value
            high = high + 1
        elif top_mean > value >= mean:
            sheet.cell(column=8, row=midhigh).value = sheet.cell(
                column=1, row=row
            ).value
            sheet.cell(column=9, row=midhigh).value = value
            midhigh = midhigh + 1
        elif mean > value >= bottom_mean:
            sheet.cell(column=12, row=midlow).value = sheet.cell(
                column=1, row=row
            ).value
            sheet.cell(column=13, row=midlow).value = value
            midlow = midlow + 1
        else:
            sheet.cell(column=16, row=low).value = sheet.cell(column=1, row=row).value
            sheet.cell(column=17, row=low).value = value
            low = low + 1


def get_stats(sheet, newsheet, max_column, max_row):
    question_data: list = [("Question", "Score")]
    for k in range(2, max_column):
        question_data.append(
            (
                sheet.cell(column=k, row=1).value,
                sheet.cell(column=k, row=max_row + 1).value,
            )
        )
    for row in question_data:
        newsheet.append(row)
    mean_values_question: dict = sort_mean_top_bottom(newsheet)
    max_row_student = sort_mean_question(
        mean=mean_values_question["mean"],
        top_mean=mean_values_question["top_mean"],
        bottom_mean=mean_values_question["bottom_mean"],
        sheet=newsheet,
    )
    new_sheet_max_row = newsheet.max_row
    for i in range(1, new_sheet_max_row + 1):
        newsheet.cell(column=1, row=i).value = None
        newsheet.cell(column=2, row=i).value = None
    student_data: list = [("Student Name", "Score")]
    for k in range(2, max_row):
        student_data.append(
            (
                sheet.cell(column=1, row=k).value,
                sheet.cell(column=max_column + 1, row=k).value,
            )
        )
    row_count = 1
    for row in student_data:
        newsheet.cell(column=1, row=row_count).value = row[0]
        newsheet.cell(column=2, row=row_count).value = row[1]
        row_count = row_count + 1
    student_mean_values: dict = sort_mean_top_bottom(newsheet)
    max_row_results = len(student_data)
    sort_mean_student(
       mean=student_mean_values["mean"],
       top_mean=student_mean_values["top_mean"],
       bottom_mean=student_mean_values["bottom_mean"],
       sheet=newsheet,
       max_row=max_row_student,
       max_row_results=max_row_results,
   )


def sort_sheet(sheet, newsheet):
    max_row = sheet.max_row
    max_column = sheet.max_column
    sort_student_scores(max_column=max_column, max_row=max_row, sheet=sheet)
    sort_question_scores(max_column=max_column, max_row=max_row, sheet=sheet)
    get_stats(sheet=sheet, newsheet=newsheet, max_column=max_column, max_row=max_row)


def create_question_chart(sheet):
    max_row = sheet.max_row
    max_column = sheet.max_column
    chart = BarChart()
    chart.type = "bar"
    chart.style = 11
    chart.title = "Questions Totals"
    chart.x_axis.title = "Questions"
    chart.y_axis.title = "Score"
    data = Reference(sheet, min_row=2, max_row=max_row, min_col=1, max_col=max_column)
    cats = Reference(sheet, min_col=1, min_row=1, max_row=max_row)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    sheet.add_chart(chart, "A{}".format(max_row + 2))


def main():
    workbook = get_workbook()
    if workbook:
        for sheet in workbook.sheetnames:
            newsheet = workbook.create_sheet("{}_stats".format(sheet))
            sort_sheet(workbook[sheet], newsheet)
        workbook.save("Sorted_Sheet.xlsx")


main()
